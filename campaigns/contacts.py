import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


MAX_CONTACTS = 1000


@dataclass(frozen=True)
class Contact:
    name: str
    phone: str

    def to_dict(self):
        return {"nome": self.name, "numero": self.phone}


class ContactFileError(ValueError):
    pass


def _normalize_header(value):
    value = str(value or "").strip().lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return re.sub(r"\s+", "", value)


def normalize_phone(value):
    raw = str(value or "").strip()
    if not raw:
        raise ContactFileError("Número vazio.")

    digits = re.sub(r"\D", "", raw)

    if digits.startswith("00"):
        digits = digits[2:]

    # Conveniência para bases brasileiras sem DDI.
    if len(digits) in (10, 11):
        digits = "55" + digits

    if not (10 <= len(digits) <= 15):
        raise ContactFileError(f"Número inválido: {raw}")

    return digits


def _find_columns(headers):
    normalized = {_normalize_header(h): idx for idx, h in enumerate(headers)}

    name_keys = ("nome", "name")
    phone_keys = ("numero", "número", "telefone", "phone", "celular", "whatsapp")

    name_idx = next((normalized[k] for k in name_keys if k in normalized), None)
    phone_idx = next((normalized[_normalize_header(k)] for k in phone_keys if _normalize_header(k) in normalized), None)

    if name_idx is None or phone_idx is None:
        raise ContactFileError(
            "A planilha precisa conter as colunas 'Nome' e 'Numero'."
        )

    return name_idx, phone_idx


def _build_contacts(rows):
    rows = list(rows)
    if not rows:
        raise ContactFileError("Arquivo sem dados.")

    headers = [str(v or "").strip() for v in rows[0]]
    name_idx, phone_idx = _find_columns(headers)

    contacts = []
    errors = []
    seen = set()

    for row_number, row in enumerate(rows[1:], start=2):
        row = list(row)

        name = str(row[name_idx] if name_idx < len(row) else "").strip()
        phone_raw = row[phone_idx] if phone_idx < len(row) else ""

        if not name and not str(phone_raw or "").strip():
            continue

        if not name:
            errors.append(f"Linha {row_number}: Nome vazio.")
            continue

        try:
            phone = normalize_phone(phone_raw)
        except ContactFileError as exc:
            errors.append(f"Linha {row_number}: {exc}")
            continue

        if phone in seen:
            continue

        seen.add(phone)
        contacts.append(Contact(name=name, phone=phone))

        if len(contacts) > MAX_CONTACTS:
            raise ContactFileError(
                f"A versão Alpha aceita até {MAX_CONTACTS} destinatários por campanha."
            )

    if not contacts:
        details = " ".join(errors[:5])
        raise ContactFileError(
            "Nenhum contato válido encontrado." + (f" {details}" if details else "")
        )

    return contacts, errors


def parse_contacts(filename, content):
    suffix = Path(filename or "").suffix.lower()

    if suffix == ".csv":
        text = None
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if text is None:
            raise ContactFileError("Não foi possível decodificar o CSV.")

        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"

        rows = list(csv.reader(io.StringIO(text), dialect))
        return _build_contacts(rows)

    if suffix == ".xlsx":
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
            return _build_contacts(rows)
        except ContactFileError:
            raise
        except Exception as exc:
            raise ContactFileError(f"Falha lendo XLSX: {exc}") from exc

    if suffix == ".xls":
        raise ContactFileError(
            "Arquivos .xls antigos não são suportados nesta versão. Salve como .xlsx ou .csv."
        )

    raise ContactFileError("Formato inválido. Use .csv ou .xlsx.")
